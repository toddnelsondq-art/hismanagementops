from __future__ import annotations

import re
from collections import OrderedDict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName


INPUT = Path("outputs/location-dropdown-workbook/DQ_Maintenance_Tracking_System_with_Location_Dropdowns_location_dropdowns.xlsx")
OUTPUT_DIR = Path("outputs/location-dropdown-workbook")
OUTPUT = OUTPUT_DIR / "DQ_Maintenance_Tracking_System_with_Location_Dropdowns_dependent_dropdowns.xlsx"


def safe_name(value: object) -> str:
    text = str(value).strip()
    text = re.sub(r"[^A-Za-z0-9_]", "_", text)
    if not text:
        text = "Blank"
    return text


def clear_validations(ws, target_ranges: set[str]) -> None:
    kept = []
    for dv in ws.data_validations.dataValidation:
        sqref = str(dv.sqref)
        if sqref not in target_ranges:
            kept.append(dv)
    ws.data_validations.dataValidation = kept


def add_list_validation(ws, cell_range: str, formula: str, prompt: str | None = None) -> None:
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.showErrorMessage = False
    if prompt:
        dv.showInputMessage = True
        dv.promptTitle = "Choose from list"
        dv.prompt = prompt
    ws.add_data_validation(dv)
    dv.add(cell_range)


def upsert_defined_name(wb, name: str, attr_text: str) -> None:
    try:
        wb.defined_names.delete(name)
    except Exception:
        pass
    wb.defined_names.add(DefinedName(name=name, attr_text=attr_text))


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = load_workbook(INPUT)

    locations = wb["Locations"]
    equipment = wb["Equipment"]
    lists = wb["Lists"]

    # Use the real Locations sheet as the source of truth for location dropdowns.
    location_rows: list[tuple[object, object]] = []
    for row in range(2, locations.max_row + 1):
        location_id = locations.cell(row, 1).value
        location_name = locations.cell(row, 2).value
        if location_id not in (None, "") and location_name not in (None, ""):
            location_rows.append((location_id, location_name))

    # Normalize existing sample equipment/work-order/PM location IDs from DQ-01 style
    # to the actual store IDs when they appear in sequence.
    dq_to_location = {
        f"DQ-{idx:02d}": location
        for idx, location in enumerate(location_rows, start=1)
    }

    def normalize_sheet_locations(sheet_name: str, id_col: int, name_col: int) -> None:
        ws = wb[sheet_name]
        for row in range(2, ws.max_row + 1):
            value = ws.cell(row, id_col).value
            if value in dq_to_location:
                location_id, location_name = dq_to_location[value]
                ws.cell(row, id_col).value = location_id
                ws.cell(row, name_col).value = location_name

    normalize_sheet_locations("Equipment", 2, 3)
    normalize_sheet_locations("Work Orders", 3, 4)
    normalize_sheet_locations("PM Schedule", 2, 3)

    # Rebuild the user-editable Lists location columns from Locations.
    lists.cell(1, 10).value = "Location ID Dropdown"
    lists.cell(1, 11).value = "Location Name Dropdown"
    lists.cell(1, 12).value = "Service Area Dropdown"
    for row in range(2, max(lists.max_row, 60) + 1):
        for col in (10, 11, 12):
            lists.cell(row, col).value = None
    for idx, (location_id, location_name) in enumerate(location_rows, start=2):
        lists.cell(idx, 10).value = location_id
        lists.cell(idx, 11).value = location_name
        lists.cell(idx + 1, 12).value = location_name
    lists.cell(2, 12).value = "All Locations"

    # Make the editable category list explicit and easy to maintain.
    existing_categories = []
    for row in range(2, lists.max_row + 1):
        value = lists.cell(row, 3).value
        if value not in (None, "") and value not in existing_categories:
            existing_categories.append(value)
    for default in ["HVAC", "Refrigeration", "Plumbing", "Electrical", "Equipment", "Building", "Signage", "Food Safety", "POS/IT", "Other"]:
        if default not in existing_categories:
            existing_categories.append(default)
    lists.cell(1, 3).value = "Category"
    for row in range(2, 80):
        lists.cell(row, 3).value = None
    for idx, value in enumerate(existing_categories, start=2):
        lists.cell(idx, 3).value = value

    # Build/update hidden helper sheet for location-specific equipment dropdowns.
    helper_name = "Dropdown Data"
    if helper_name in wb.sheetnames:
        helper = wb[helper_name]
        helper.delete_rows(1, helper.max_row)
    else:
        helper = wb.create_sheet(helper_name)
    helper.sheet_state = "hidden"

    equipment_by_location: OrderedDict[object, list[tuple[object, object]]] = OrderedDict()
    for location_id, _location_name in location_rows:
        equipment_by_location[location_id] = []
    for row in range(2, equipment.max_row + 1):
        equipment_id = equipment.cell(row, 1).value
        location_id = equipment.cell(row, 2).value
        equipment_name = equipment.cell(row, 4).value
        if location_id in equipment_by_location and equipment_id not in (None, ""):
            equipment_by_location[location_id].append((equipment_id, equipment_name or equipment_id))

    col = 1
    max_items = 1
    for location_id, location_name in location_rows:
        id_col = col
        name_col = col + 1
        helper.cell(1, id_col).value = f"Equipment IDs - {location_name}"
        helper.cell(1, name_col).value = f"Equipment Names - {location_name}"
        helper.cell(1, id_col).font = Font(bold=True)
        helper.cell(1, name_col).font = Font(bold=True)
        items = equipment_by_location.get(location_id, [])
        if not items:
            items = [("", "")]
        max_items = max(max_items, len(items))
        for row_idx, (equipment_id, equipment_name) in enumerate(items, start=2):
            helper.cell(row_idx, id_col).value = equipment_id
            helper.cell(row_idx, name_col).value = equipment_name
        id_letter = get_column_letter(id_col)
        name_letter = get_column_letter(name_col)
        safe = safe_name(location_id)
        last_row = max(2, len(items) + 1)
        upsert_defined_name(wb, f"EquipIDs_{safe}", f"{quote_sheetname(helper_name)}!${id_letter}$2:${id_letter}${last_row}")
        upsert_defined_name(wb, f"EquipNames_{safe}", f"{quote_sheetname(helper_name)}!${name_letter}$2:${name_letter}${last_row}")
        col += 2

    # Add a short instruction note on Lists so future edits are obvious.
    lists.cell(1, 14).value = "Dropdown Notes"
    lists.cell(2, 14).value = "Edit Categories in column C. Edit locations on the Locations tab. Assign equipment to a location on the Equipment tab."
    lists.cell(3, 14).value = "Work Orders and PM Schedule equipment choices depend on the Location ID selected in that same row."
    lists.cell(1, 14).font = Font(bold=True)
    lists.column_dimensions["N"].width = 95

    last_location_row = len(location_rows) + 1
    last_service_area_row = len(location_rows) + 2
    last_category_row = len(existing_categories) + 1

    location_id_formula = f"=Lists!$J$2:$J${last_location_row}"
    location_name_formula = f"=Lists!$K$2:$K${last_location_row}"
    service_area_formula = f"=Lists!$L$2:$L${last_service_area_row}"
    category_formula = f"=Lists!$C$2:$C${last_category_row}"
    equipment_type_formula = "=Lists!$D$2:$D$12"

    # Remove/replace validations that we own.
    clear_validations(wb["Work Orders"], {"C2:C500", "D2:D500", "F2:F500", "G2:G500", "H2:H500"})
    clear_validations(wb["Equipment"], {"B2:B500", "C2:C500", "E2:E500"})
    clear_validations(wb["PM Schedule"], {"B2:B500", "C2:C500", "D2:D500", "E2:E500"})
    clear_validations(wb["Vendors"], {"C2:C500", "H2:H500"})

    work_orders = wb["Work Orders"]
    add_list_validation(work_orders, "C2:C500", location_id_formula, "Choose a location first.")
    add_list_validation(work_orders, "D2:D500", location_name_formula)
    add_list_validation(work_orders, "F2:F500", category_formula, "Choose from the editable Category list on the Lists tab.")
    add_list_validation(work_orders, "G2:G500", '=INDIRECT("EquipIDs_"&SUBSTITUTE($C2,"-","_"))', "Choose equipment after choosing the location.")
    add_list_validation(work_orders, "H2:H500", '=INDIRECT("EquipNames_"&SUBSTITUTE($C2,"-","_"))')
    for row in range(2, 501):
        work_orders.cell(row, 4).value = f'=IFERROR(INDEX(Locations!$B$2:$B${last_location_row},MATCH(C{row},Locations!$A$2:$A${last_location_row},0)),"")'
        work_orders.cell(row, 8).value = f'=IFERROR(INDEX(Equipment!$D$2:$D$500,MATCH(G{row},Equipment!$A$2:$A$500,0)),"")'

    equipment_ws = wb["Equipment"]
    add_list_validation(equipment_ws, "B2:B500", location_id_formula, "Choose the location this equipment belongs to.")
    add_list_validation(equipment_ws, "C2:C500", location_name_formula)
    add_list_validation(equipment_ws, "E2:E500", equipment_type_formula, "Choose from the editable Equipment Type list on the Lists tab.")
    for row in range(2, 501):
        equipment_ws.cell(row, 3).value = f'=IFERROR(INDEX(Locations!$B$2:$B${last_location_row},MATCH(B{row},Locations!$A$2:$A${last_location_row},0)),"")'

    pm = wb["PM Schedule"]
    add_list_validation(pm, "B2:B500", location_id_formula, "Choose a location first.")
    add_list_validation(pm, "C2:C500", location_name_formula)
    add_list_validation(pm, "D2:D500", '=INDIRECT("EquipIDs_"&SUBSTITUTE($B2,"-","_"))', "Choose equipment after choosing the location.")
    add_list_validation(pm, "E2:E500", '=INDIRECT("EquipNames_"&SUBSTITUTE($B2,"-","_"))')
    for row in range(2, 501):
        pm.cell(row, 3).value = f'=IFERROR(INDEX(Locations!$B$2:$B${last_location_row},MATCH(B{row},Locations!$A$2:$A${last_location_row},0)),"")'
        pm.cell(row, 5).value = f'=IFERROR(INDEX(Equipment!$D$2:$D$500,MATCH(D{row},Equipment!$A$2:$A$500,0)),"")'

    vendors = wb["Vendors"]
    add_list_validation(vendors, "C2:C500", category_formula, "Choose from the editable Category list on the Lists tab.")
    add_list_validation(vendors, "H2:H500", service_area_formula)

    # Make the important editable list columns readable.
    for ws in [lists, helper]:
        ws.freeze_panes = "A2"
    lists.column_dimensions["C"].width = 20
    lists.column_dimensions["J"].width = 18
    lists.column_dimensions["K"].width = 24
    lists.column_dimensions["L"].width = 24

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx in [3, 10, 11, 12, 14]:
        cell = lists.cell(1, col_idx)
        cell.fill = header_fill
        cell.font = header_font

    # Highlight if someone picks equipment before choosing a location.
    warning_fill = PatternFill("solid", fgColor="FFF2CC")
    for ws, loc_col, equip_col in [(work_orders, "C", "G"), (pm, "B", "D")]:
        rng = f"{equip_col}2:{equip_col}500"
        ws.conditional_formatting.add(
            rng,
            FormulaRule(formula=[f'=AND(${equip_col}2<>"",${loc_col}2="")'], fill=warning_fill),
        )

    wb.save(OUTPUT)


if __name__ == "__main__":
    main()
