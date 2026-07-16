from pathlib import Path
from openpyxl import load_workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import PatternFill, Font, Border, Side

input_path = Path(r"C:\Users\tanel\OneDrive\Escritorio\DQ_Maintenance_Tracking_System_with_Location_Dropdowns.xlsx")
output_dir = Path(r"C:\Users\tanel\Documents\Codex\2026-07-02\i\outputs\location-dropdown-workbook")
output_dir.mkdir(parents=True, exist_ok=True)
output_path = output_dir / "DQ_Maintenance_Tracking_System_with_Location_Dropdowns_location_dropdowns.xlsx"

wb = load_workbook(input_path)

locations_ws = wb["Locations"]
lists_ws = wb["Lists"]

locations = []
for row in locations_ws.iter_rows(min_row=2, max_row=locations_ws.max_row, min_col=1, max_col=2, values_only=True):
    location_id, location_name = row
    if location_id not in (None, "") and location_name not in (None, ""):
        locations.append((location_id, location_name))

for row in range(1, 250):
    for col in range(10, 13):
        lists_ws.cell(row=row, column=col).value = None

lists_ws["J1"] = "Location ID Dropdown"
lists_ws["K1"] = "Location Name Dropdown"
lists_ws["L1"] = "Service Area Dropdown"

for idx, (location_id, location_name) in enumerate(locations, start=2):
    lists_ws.cell(row=idx, column=10).value = location_id
    lists_ws.cell(row=idx, column=11).value = location_name
    lists_ws.cell(row=idx + 1, column=12).value = location_name
lists_ws["L2"] = "All Locations"

header_fill = PatternFill("solid", fgColor="0A2E5C")
header_font = Font(bold=True, color="FFFFFF")
thin = Side(style="thin", color="DCE4EE")
for cell in lists_ws["J1:L1"][0]:
    cell.fill = header_fill
    cell.font = header_font
for row in lists_ws.iter_rows(min_row=1, max_row=max(len(locations) + 2, 2), min_col=10, max_col=12):
    for cell in row:
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

id_formula = f"=Lists!$J$2:$J${len(locations) + 1}"
name_formula = f"=Lists!$K$2:$K${len(locations) + 1}"
service_formula = f"=Lists!$L$2:$L${len(locations) + 2}"

targets = [
    ("Work Orders", "C2:C500", id_formula),
    ("Work Orders", "D2:D500", name_formula),
    ("Equipment", "B2:B500", id_formula),
    ("Equipment", "C2:C500", name_formula),
    ("PM Schedule", "B2:B500", id_formula),
    ("PM Schedule", "C2:C500", name_formula),
    ("Vendors", "H2:H500", service_formula),
]

for sheet_name, cell_range, formula in targets:
    ws = wb[sheet_name]
    dv = DataValidation(type="list", formula1=formula, allow_blank=True)
    dv.error = "Choose a value from the dropdown list."
    dv.errorTitle = "Invalid selection"
    dv.prompt = "Choose from the location list."
    dv.promptTitle = "Location dropdown"
    ws.add_data_validation(dv)
    dv.add(cell_range)

for sheet_name in ["Work Orders", "Equipment", "PM Schedule", "Vendors", "Lists"]:
    wb[sheet_name].freeze_panes = "A2"

wb.save(output_path)
print(output_path)
